import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
import xlsxwriter


class XLSXParser:
    def write(self, filename, ts_suffix, fault_suffix, apfd_value, tc_order):
        writer = pd.ExcelWriter('../resources/xlsx/' +
                                filename + '.xlsx', engine="openpyxl")
        startrow = None
        sheet_name = filename + fault_suffix + ts_suffix
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError

        try:
            writer.book = load_workbook('../resources/xlsx/' +
                                        filename + '.xlsx')
            if sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass
        df_array = [[tc_order, apfd_value]]
        df = pd.DataFrame(df_array)
        if startrow is None:
            startrow = 0
            header_array = [["TC Order", "APFD"]]
            header_array.extend(df_array)
            df = pd.DataFrame(header_array)
        df.to_excel(writer, sheet_name, startrow=startrow, header=None)
        # save the workbook
        writer.save()
